# ... (All imports and configuration are the same) ...
import asyncio, json, pandas as pd, numpy as np, smtplib, aiohttp, time, os, random
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, jsonify
from threading import Thread, Lock
from datetime import datetime
import pandas_ta as ta
from openpyxl import Workbook, load_workbook
import gunicorn # Keep this import

# --- (Configuration is the same ) ---
TOP_N_COINS = 20
RSI_PERIOD = 14
EMA_PERIOD = 20
RSI_THRESHOLD = 80
TIMEFRAME = '1h'
CALIBRATION_OFFSET = 3
COOLDOWN_PERIOD = 2 * 60 * 60
DATA_CACHE_FILE = 'initial_data.json'
ALERT_COUNTER_FILE = 'alert_counter.json'
DATABASE_FILE = 'alerts_database.xlsx'
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_SENDER = os.environ.get('EMAIL_SENDER')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD')
EMAIL_RECEIVER = os.environ.get('EMAIL_RECEIVER')
app_state = {"market_data": {}, "alert_log": [], "alerted_coins": {}, "init_progress": 0, "total_symbols": 0, "rsi_over_80_count": 0}
db_lock = Lock()

# --- (IndicatorCalculator and helper functions are UNCHANGED) ---
class IndicatorCalculator:
    def __init__(self, symbol, rsi_period=14, ema_period=20):
        self.symbol = symbol; self.rsi_period = rsi_period; self.ema_period = ema_period
        self.df = pd.DataFrame(columns=['open', 'high', 'low', 'close', 'volume'])
        self.last_indicators = {}
    def initialize_with_data(self, klines):
        if not isinstance(klines, list) or not all(isinstance(row, list) for row in klines): return False
        self.df = pd.DataFrame(klines, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'close_time', 'quote_asset_volume', 'number_of_trades', 'taker_buy_base_asset_volume', 'taker_buy_quote_asset_volume', 'ignore'], dtype=float)
        self._perform_calculation(); return True
    async def initialize_from_api(self):
        url = f"https://fapi.binance.com/fapi/v1/klines?symbol={self.symbol}&interval={TIMEFRAME}&limit=500"
        try:
            async with aiohttp.ClientSession( ) as session:
                async with session.get(url) as response:
                    if response.status == 200:
                        klines = await response.json()
                        self.df = pd.DataFrame(klines[:-1], columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'close_time', 'quote_asset_volume', 'number_of_trades', 'taker_buy_base_asset_volume', 'taker_buy_quote_asset_volume', 'ignore'], dtype=float)
                        self._perform_calculation(); return klines[:-1]
                    return None
        except Exception: return None
    def calculate_live_indicators(self, live_price):
        if self.df.empty: return
        live_df = self.df.copy(); live_df.loc[live_df.index[-1], 'close'] = live_price
        self._perform_calculation(df=live_df)
    def add_closed_candle(self, kline_data):
        new_row = pd.DataFrame([kline_data], columns=self.df.columns, dtype=float)
        self.df = pd.concat([self.df, new_row], ignore_index=True)
        if len(self.df) > 500: self.df = self.df.iloc[1:]
        return self._perform_calculation()
    def _perform_calculation(self, df=None):
        target_df = df if df is not None else self.df
        if len(target_df) < self.rsi_period: return None
        rsi = ta.rsi(close=target_df['close'], length=self.rsi_period)
        macd = ta.macd(close=target_df['close'])
        ema = ta.ema(close=target_df['close'], length=self.ema_period)
        if rsi is None or macd is None or ema is None or rsi.empty or macd.empty or ema.empty: return None
        macd_col_name = [col for col in macd.columns if col.startswith('MACD_')][0]
        macds_col_name = [col for col in macd.columns if col.startswith('MACDs_')][0]
        if len(target_df) > 24:
            price_24h_ago = target_df['close'].iloc[-25]
            current_price = target_df['close'].iloc[-1]
            change_24h = ((current_price - price_24h_ago) / price_24h_ago) * 100 if price_24h_ago != 0 else 0
        else: change_24h = 0
        self.last_indicators = {
            'price': target_df['close'].iloc[-1], 'rsi': rsi.iloc[-1] - CALIBRATION_OFFSET,
            'macd': macd[macd_col_name].iloc[-1], 'macds': macd[macds_col_name].iloc[-1],
            'ema': ema.iloc[-1], 'change_24h': change_24h
        }
        app_state["market_data"][self.symbol] = self.last_indicators
        return self.last_indicators

# --- (All helper functions like write_to_database, send_email_alert, etc. are UNCHANGED) ---
def write_to_database(alert_data):
    with db_lock:
        try:
            file_exists = os.path.isfile(DATABASE_FILE)
            if not file_exists:
                workbook = Workbook(); sheet = workbook.active; sheet.title = "Alerts"
                sheet.append(["Alert #", "Timestamp", "Symbol", "Price", "RSI", "MACD", "MACD Signal", "EMA", "24h Change %"])
                workbook.save(DATABASE_FILE)
            workbook = load_workbook(DATABASE_FILE); sheet = workbook.active
            sheet.append([alert_data['alert_num'], alert_data['time'], alert_data['symbol'], f"{alert_data['price']:.4f}", f"{alert_data['rsi']:.2f}", f"{alert_data['macd']:.4f}", f"{alert_data['macds']:.4f}", f"{alert_data['ema']:.4f}", f"{alert_data['change_24h']:.2f}%"])
            workbook.save(DATABASE_FILE)
            print(f"💾 Alert #{alert_data['alert_num']} successfully saved to {DATABASE_FILE}")
        except Exception as e: print(f"🚨 CRITICAL: Could not write to database file! Error: {e}")
def get_next_alert_number():
    alert_num = 0
    try:
        if os.path.exists(ALERT_COUNTER_FILE):
            with open(ALERT_COUNTER_FILE, 'r') as f: data = json.load(f)
            alert_num = data.get('last_alert_number', 0)
    except (IOError, json.JSONDecodeError): alert_num = 0
    next_alert_num = alert_num + 1
    try:
        with open(ALERT_COUNTER_FILE, 'w') as f: json.dump({'last_alert_number': next_alert_num}, f)
    except IOError as e: print(f"🚨 CRITICAL: Could not write to alert counter file! {e}")
    return next_alert_num
def send_email_alert(symbol, indicators):
    try:
        alert_number = get_next_alert_number()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        alert_data = {'alert_num': alert_number, 'time': now, 'symbol': symbol, **indicators}
        html_body = f"""<html><head><style>body {{ font-family: sans-serif; }} table {{ border-collapse: collapse; width: 100%; max-width: 600px; }} th, td {{ border: 1px solid #dddddd; text-align: left; padding: 8px; }} th {{ background-color: #f2f2f2; }} .title {{ font-size: 1.2em; font-weight: bold; }}</style></head><body><p class="title">Alert #{alert_number}: {symbol} has crossed the RSI threshold!</p><table><tr><th>Metric</th><th>Value</th></tr><tr><td>Timestamp</td><td>{alert_data['time']}</td></tr><tr><td>Symbol</td><td>{alert_data['symbol']}</td></tr><tr><td>Price</td><td>${alert_data['price']:.4f}</td></tr><tr><td><b>RSI (Calibrated)</b></td><td><b>{alert_data['rsi']:.2f}</b></td></tr><tr><td>MACD</td><td>{alert_data['macd']:.4f}</td></tr><tr><td>MACD Signal</td><td>{alert_data['macds']:.4f}</td></tr><tr><td>EMA ({EMA_PERIOD}-period)</td><td>${alert_data['ema']:.4f}</td></tr><tr><td>24h Change</td><td>{alert_data['change_24h']:.2f}%</td></tr></table></body></html>"""
        msg = MIMEMultipart(); msg['Subject'] = f"Alert: {alert_number} - {symbol} - {RSI_THRESHOLD}"; msg['From'] = EMAIL_SENDER; msg['To'] = EMAIL_RECEIVER
        msg.attach(MIMEText(html_body, 'html'))
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls(); server.login(EMAIL_SENDER, EMAIL_PASSWORD); server.send_message(msg)
        app_state['alert_log'].insert(0, {'alert_num': alert_number, 'symbol': symbol, 'rsi': alert_data['rsi'], 'time': now})
        if len(app_state['alert_log']) > 50: app_state['alert_log'].pop()
        write_to_database(alert_data)
        print(f"✅ Email for Alert #{alert_number} sent!")
    except Exception as e: print(f"🚨 Failed to send email for {symbol}: {e}")
def check_rsi_and_alert(symbol, indicators):
    if not indicators: return
    rsi_value = indicators.get('rsi')
    if rsi_value is None: return
    update_rsi_over_80_count()
    if rsi_value > RSI_THRESHOLD:
        if symbol not in app_state["alerted_coins"] or (time.time() - app_state["alerted_coins"][symbol]) > COOLDOWN_PERIOD:
            print(f"--- ALERT TRIGGERED for {symbol} (Calibrated RSI: {rsi_value}) ---")
            send_email_alert(symbol, indicators)
            app_state["alerted_coins"][symbol] = time.time()
async def get_all_futures_pairs():
    url = "https://fapi.binance.com/fapi/v1/exchangeInfo"
    try:
        async with aiohttp.ClientSession( ) as session:
            async with session.get(url) as response:
                data = await response.json()
                return [s['symbol'] for s in data['symbols'] if s['quoteAsset'] == 'USDT' and s['contractType'] == 'PERPETUAL' and s['status'] == 'TRADING']
    except Exception: return []
def update_rsi_over_80_count():
    app_state["rsi_over_80_count"] = sum(1 for data in app_state["market_data"].values() if data.get('rsi', 0) > RSI_THRESHOLD)

async def binance_websocket_listener():
    print("Starting Binance listener...")
    symbols = await get_all_futures_pairs()
    if not symbols: print("Could not fetch symbol list. Exiting."); return
    app_state["total_symbols"] = len(symbols)
    calculators = {symbol: IndicatorCalculator(symbol) for symbol in symbols}
    
    cache_is_valid = os.path.exists(DATA_CACHE_FILE)
    if cache_is_valid:
        print("Loading initial data from cache...")
        with open(DATA_CACHE_FILE, 'r') as f: cached_data = json.load(f)
        for i, symbol in enumerate(symbols):
            if not (symbol in cached_data and calculators[symbol].initialize_with_data(cached_data[symbol])):
                cache_is_valid = False; print(f"Stale or invalid cache detected. Re-fetching all data."); break
            app_state["init_progress"] = int(((i + 1) / len(symbols)) * 100)
    
    if not cache_is_valid:
        app_state["init_progress"] = 0; print("Fetching initial data from API...")
        new_cache = {}
        # --- FIX #1: Limit simultaneous requests ---
        sem = asyncio.Semaphore(10) # Only allow 10 requests at a time
        async def fetch_and_init(symbol):
            async with sem:
                historical_data = await calculators[symbol].initialize_from_api()
                if historical_data: new_cache[symbol] = historical_data
                with db_lock:
                    progress = int((len(new_cache) / len(symbols)) * 100)
                    if progress > app_state["init_progress"]: app_state["init_progress"] = progress
        await asyncio.gather(*(fetch_and_init(s) for s in symbols))
        print("Saving new data to cache file...");
        with open(DATA_CACHE_FILE, 'w') as f: json.dump(new_cache, f)
        app_state["init_progress"] = 100

    print("\n--- Initial checks complete. Connecting to WebSockets. ---")
    for symbol, calc in calculators.items():
        if calc.last_indicators: check_rsi_and_alert(symbol, calc.last_indicators)
    
    async def listen_for_closed_candles():
        # ... (This function is unchanged and correct)
        streams = [f"{s.lower()}@kline_{TIMEFRAME}" for s in symbols]
        chunk_size = 100
        async def listen_chunk(stream_chunk):
            websocket_url = f"wss://fstream.binance.com/stream?streams={'/'.join(stream_chunk)}"
            while True:
                try:
                    async with aiohttp.ClientSession( ) as session:
                        async with session.ws_connect(websocket_url) as ws:
                            print(f"Official {TIMEFRAME} Candle Listener connected for chunk.")
                            async for msg in ws:
                                data = json.loads(msg.data)
                                if 'k' in data['data'] and data['data']['k']['x']:
                                    kline = data['data']['k']; symbol = kline['s']; calc = calculators.get(symbol)
                                    if calc:
                                        kline_data_for_df = [kline['t'], kline['o'], kline['h'], kline['l'], kline['c'], kline['v'], kline['T'], kline['q'], kline['n'], kline['V'], kline['Q'], kline['B']]
                                        indicators = calc.add_closed_candle(kline_data_for_df)
                                        check_rsi_and_alert(symbol, indicators)
                except (aiohttp.ClientError, asyncio.TimeoutError, Exception ) as e:
                    print(f"1h Listener Error: {e}. Reconnecting in 10-15 seconds...")
                    await asyncio.sleep(10 + random.uniform(0, 5))
        tasks = [listen_chunk(streams[i:i + chunk_size]) for i in range(0, len(streams), chunk_size)]
        await asyncio.gather(*tasks)

    async def listen_for_top_n_tickers():
        # ... (This function is mostly the same, but with the timeout fix)
        current_subscriptions = set()
        while True:
            try:
                sorted_by_rsi = sorted([s for s in app_state["market_data"].items() if s[1].get('rsi')], key=lambda item: item[1]['rsi'], reverse=True)
                top_n_symbols = {s[0] for s in sorted_by_rsi[:TOP_N_COINS]}
                if top_n_symbols != current_subscriptions:
                    print(f"[Ticker Manager] Updating subscriptions. New Top {TOP_N_COINS}: {len(top_n_symbols)} coins.")
                    current_subscriptions = top_n_symbols
                if not current_subscriptions:
                    await asyncio.sleep(5); continue
                streams = [f"{s.lower()}@ticker" for s in current_subscriptions]
                websocket_url = f"wss://fstream.binance.com/stream?streams={'/'.join(streams)}"
                async with aiohttp.ClientSession( ) as session:
                    async with session.ws_connect(websocket_url) as ws:
                        while True:
                            try:
                                # --- FIX #2: Increase the timeout ---
                                msg = await asyncio.wait_for(ws.receive(), timeout=30.0) # Increased to 30 seconds
                                if msg.type == aiohttp.WSMsgType.TEXT:
                                    data = json.loads(msg.data )['data']
                                    symbol = data['s']; calc = calculators.get(symbol)
                                    if calc: calc.calculate_live_indicators(float(data['c']))
                                elif msg.type in (aiohttp.WSMsgType.CLOSED, aiohttp.WSMsgType.ERROR ):
                                    print("Dynamic Ticker connection closed/errored. Breaking to reconnect."); break
                            except asyncio.TimeoutError:
                                print("[Ticker Manager] 30s timeout reached. Re-evaluating Top 20 coins."); break
                update_rsi_over_80_count()
            except (aiohttp.ClientError, asyncio.TimeoutError, Exception ) as e:
                print(f"Dynamic Ticker Main Loop Error: {e}. Reconnecting in 10-15 seconds...")
                await asyncio.sleep(10 + random.uniform(0, 5))

    await asyncio.gather(listen_for_closed_candles(), listen_for_top_n_tickers())

# --- (Flask server part is UNCHANGED) ---
app = Flask(__name__)
# ... (all the @app.route functions are the same)
@app.route('/')
def index(): return app.send_static_file('index.html')
@app.route('/data')
def get_data():
    market_list = []
    for symbol, data in app_state["market_data"].items():
        if data and 'rsi' in data: market_list.append({'symbol': symbol, **data})
    sorted_market_data = sorted(market_list, key=lambda x: x.get('rsi', 0), reverse=True)
    enriched_alert_log = []
    for alert in app_state["alert_log"]:
        symbol = alert['symbol']; current_data = app_state["market_data"].get(symbol)
        current_rsi = current_data.get('rsi') if current_data else 'N/A'
        enriched_alert_log.append({'alert_num': alert.get('alert_num', '-'), 'time': alert['time'], 'symbol': symbol, 'sent_rsi': alert['rsi'], 'live_rsi': current_rsi})
    response_data = {"market_data": sorted_market_data, "alert_log": enriched_alert_log, "init_progress": app_state["init_progress"], "rsi_over_80_count": app_state["rsi_over_80_count"]}
    return jsonify(response_data)
@app.route('/database')
def get_database():
    db_data = []
    if os.path.exists(DATABASE_FILE):
        with db_lock:
            try:
                df = pd.read_excel(DATABASE_FILE)
                db_data = df.to_dict('records')
            except Exception as e:
                print(f"Error reading database file: {e}")
                return jsonify({"error": "Could not read database file."}), 500
    enriched_db = []
    for record in db_data:
        symbol = record.get('Symbol')
        live_data = app_state['market_data'].get(symbol)
        if live_data: record['live_price'] = live_data.get('price')
        enriched_db.append(record)
    return jsonify(enriched_db)

# --- This part is for Railway to start the app ---
binance_thread = Thread(target=run_asyncio_loop, daemon=True)
binance_thread.start()
